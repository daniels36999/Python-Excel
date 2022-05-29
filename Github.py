import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference,  LineChart
from git import Repo

#POCISIONES TABLA
FilInico=4 
ColDias=1
ColTemp=2
ColPeso=3
ColAlt=4
ColIMC=5
ColO2Sat=6

#LECTURA ARCHIVO EXCEL
wb=load_workbook('ejemplo.xlsx')
HojaExcel=wb['hoja1']

# del HojaExcel._charts[1]
# del HojaExcel._charts[2]
# del HojaExcel._charts[3]
# del HojaExcel._charts[4]
# del HojaExcel._charts[5]

#ASIGNACION DE DIMENCION DE LOS DATOS
min_col=wb.active.min_column
max_col=wb.active.max_column
min_fil=wb.active.min_row
max_fil=wb.active.max_row

#POCISIONES GRAFICAS
PosTemp='A'+str(max_fil+2)
PosPeso='A'+str(max_fil+17)
PosAlt='A'+str(max_fil+32)
PosIMC='A'+str(max_fil+47)
PosO2Sat='A'+str(max_fil+62)

#CONDICIONES PARA LAS GRAFICAS TEMPERATURA CORPORAL
GrafTemp= LineChart()
DTTemp=Reference(HojaExcel,min_col=ColTemp,max_col=ColTemp,min_row=FilInico,max_row=max_fil)
DTDia=Reference(HojaExcel,min_col=ColDias,max_col=ColDias,min_row=FilInico,max_row=max_fil)
GrafTemp.add_data(DTTemp,titles_from_data=True)
GrafTemp.set_categories(DTDia)
HojaExcel.add_chart(GrafTemp,PosTemp)
GrafTemp.y_axis.title = "Temperatura Corporal [°C]"
GrafTemp.x_axis.title = "Dias"
GrafTemp.title='TEMPERATURA CORPORAL'
GrafTemp.style=34

#CONDICIONES PARA LAS GRAFICAS PESO
GrafPeso= LineChart()
DTPes=Reference(HojaExcel,min_col=ColPeso,max_col=ColPeso,min_row=FilInico,max_row=max_fil)
GrafPeso.add_data(DTPes,titles_from_data=True)
GrafPeso.set_categories(DTDia)
HojaExcel.add_chart(GrafPeso,PosPeso)
GrafPeso.y_axis.title = "Peso Corporal[Kg]"
GrafPeso.x_axis.title = "Dias"
GrafPeso.title='PESO CORPORAL'
GrafPeso.style=6

#CONDICIONES PARA LAS GRAFICAS ALTURA
GrafAlt= LineChart()
DTAlt=Reference(HojaExcel,min_col=ColAlt,max_col=ColAlt,min_row=FilInico,max_row=max_fil)
GrafAlt.add_data(DTAlt,titles_from_data=True)
GrafAlt.set_categories(DTDia)
HojaExcel.add_chart(GrafAlt,PosAlt)
GrafAlt.y_axis.title = "Altura Corporal[m]"
GrafAlt.x_axis.title = "Dias"
GrafAlt.title='ALTURA CORPORAL'
GrafAlt.style=3

#CONDICIONES PARA LAS GRAFICAS IMC
GrafIMC= LineChart()
DTIMC=Reference(HojaExcel,min_col=ColIMC,max_col=ColIMC,min_row=FilInico,max_row=max_fil)
GrafIMC.add_data(DTIMC,titles_from_data=True)
GrafIMC.set_categories(DTDia)
HojaExcel.add_chart(GrafIMC,PosIMC)
GrafIMC.y_axis.title = "IMC [Kg/m^2]]"
GrafIMC.x_axis.title = "Dias"
GrafIMC.title='INDICE DE MASA CORPORAL'
GrafIMC.style=7

#CONDICIONES PARA LAS GRAFICAS O2SAT
GrafO2Sat= LineChart()
DTO2Sat=Reference(HojaExcel,min_col=ColO2Sat,max_col=ColO2Sat,min_row=FilInico,max_row=max_fil)
GrafO2Sat.add_data(DTO2Sat,titles_from_data=True)
GrafO2Sat.set_categories(DTDia)
HojaExcel.add_chart(GrafO2Sat,PosO2Sat)
GrafO2Sat.y_axis.title = "O2Sat [%]"
GrafO2Sat.x_axis.title = "Dias"
GrafO2Sat.title='SATURACION DE OXIGENO EN LA SANGRE'
GrafO2Sat.style=5


wb.save('ejemplo.xlsx')

repo = Repo('/home/pi/Desktop/GitHub1/Python-Excel')  # if repo is CWD just do '.'

repo.index.add(['ejemplo.xlsx'])
repo.index.commit('Subida1')
origin = repo.remote('origin')
origin.push()